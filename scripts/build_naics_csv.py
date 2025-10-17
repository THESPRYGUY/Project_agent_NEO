"""
Builds NAICS 2022 CSV fallback from the official Excel workbook.

- Downloads: https://www.census.gov/naics/2022NAICS/2-6%20digit_2022_Codes.xlsx
- Writes: data/naics/2-6 digit_2022_Codes.csv

Run:
  python scripts/build_naics_csv.py
"""

from __future__ import annotations

import csv
import sys
import urllib.request
from pathlib import Path


EXCEL_URL = "https://www.census.gov/naics/2022NAICS/2-6%20digit_2022_Codes.xlsx"


def ensure_openpyxl() -> None:
    try:
        import openpyxl  # noqa: F401
    except Exception:  # pragma: no cover - utility script path
        print("Installing openpyxl ...", file=sys.stderr)
        import subprocess

        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])  # nosec


def download_excel(dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    print(f"Downloading: {EXCEL_URL}\n  -> {dest}")
    req = urllib.request.Request(
        EXCEL_URL,
        headers={
            # Some hosts block default Python UA; pretend to be a browser.
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
            )
        },
    )
    with urllib.request.urlopen(req) as resp, dest.open("wb") as fh:  # nosec
        fh.write(resp.read())


def convert_excel_to_csv(xlsx_path: Path, csv_path: Path) -> int:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active

    # Find a header row (first non-empty row).
    header = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row and any(cell is not None and str(cell).strip() for cell in row):
            header = [str(cell).strip() if cell is not None else "" for cell in row]
            break
    if not header:
        raise RuntimeError("No header row found in NAICS workbook")

    lower = [h.lower() for h in header]
    code_idx = next((i for i, h in enumerate(lower) if "code" in h), None)
    title_idx = next((i for i, h in enumerate(lower) if "title" in h), None)
    if code_idx is None or title_idx is None:
        # Fallback to first two columns
        code_idx = 0
        title_idx = 1 if len(header) > 1 else 0

    rows_out: list[tuple[str, str]] = []

    def expand_code(raw: str) -> list[str]:
        s = raw.strip()
        # Handle ranges like "31-33" (Manufacturing) and similar "44-45", "48-49".
        if "-" in s:
            parts = s.split("-")
            if len(parts) == 2 and all(p.isdigit() and len(p) == 2 for p in parts):
                start, end = map(int, parts)
                if start <= end and start >= 10 and end <= 99:
                    return [f"{n:02d}" for n in range(start, end + 1)]
        # Strip trailing .0 and non-digits
        if s.endswith(".0"):
            s = s[:-2]
        s = "".join(ch for ch in s if ch.isdigit())
        return [s] if s else []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        code = (
            str(row[code_idx]).strip() if code_idx < len(row) and row[code_idx] is not None else ""
        )
        title = (
            str(row[title_idx]).strip() if title_idx < len(row) and row[title_idx] is not None else ""
        )
        if not code or not title:
            continue
        codes = expand_code(code)
        for c in codes:
            rows_out.append((c, title))

    # Deduplicate and sort
    seen: set[str] = set()
    rows_dedup: list[tuple[str, str]] = []
    for code, title in rows_out:
        if code in seen:
            continue
        seen.add(code)
        rows_dedup.append((code, title))
    rows_dedup.sort(key=lambda t: (len(t[0]), t[0]))

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["code", "title"])
        w.writerows(rows_dedup)

    return len(rows_dedup)


def main() -> int:
    repo_root = Path(__file__).resolve().parents[1]
    data_dir = repo_root / "data" / "naics"
    xlsx_path = data_dir / "2-6 digit_2022_Codes.xlsx"
    csv_path = data_dir / "2-6 digit_2022_Codes.csv"

    ensure_openpyxl()

    if not xlsx_path.exists():
        download_excel(xlsx_path)

    count = convert_excel_to_csv(xlsx_path, csv_path)
    print(f"Wrote {count} rows -> {csv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
